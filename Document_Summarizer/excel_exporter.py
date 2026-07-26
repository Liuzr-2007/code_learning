# -*- coding: utf-8 -*-
"""Excel 导出模块：将人员信息写入带样式的 Excel"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import TABLE_COLUMNS
from utils import display_width


def export_excel(save_path: str, data_list: list[dict[str, str]],
                 columns: list[str] = TABLE_COLUMNS) -> None:
    """导出汇总 Excel，中文列宽自适应

    columns 指定输出列及顺序；data_list 中每条记录用 .get(col, "") 安全取值。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "人员信息汇总"

    headers = columns
    ws.append(headers)

    # 表头样式
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    head_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 写入数据行（.get 安全取值，缺失列填空）
    for row_data in data_list:
        row = [row_data.get(col, "") for col in columns]
        ws.append(row)

    # 为数据行添加边框
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border

    # 自动列宽（使用 display_width 适应中文字符）
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, display_width(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4

    wb.save(save_path)
    print(f"\nExcel文件已保存至：{save_path}")
